import os
from openpyxl import load_workbook
from openpyxl import Workbook

filename='res/local/dblp5-local.xlsx'
if not os.path.exists(filename):
    workbook = Workbook()
    workbook.save(filename)
# 打开现有的 Excel 文件
workbook = load_workbook(filename=filename)

# 获取指定的工作表
sheet = workbook['Sheet']

# 列表数据
mode = "fedavg"
acc = [1,2,3,4]

data = [[mode]] + [[a] for a in acc]

# 获取已使用的列数
used_columns = sheet.max_column

# 写入列表数据到新的列
for row_index, row_data in enumerate(data, start=1):
    for col_index, value in enumerate(row_data, start=used_columns + 1):
        sheet.cell(row=row_index, column=col_index, value=value)

# 保存修改后的 Excel 文件
workbook.save(filename=filename)

